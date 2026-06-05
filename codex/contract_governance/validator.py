"""Validation gates for generated governance artifacts."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook

from .builder import workbook_payload
from .io import content_hash
from .schema import Audience, GovernanceSSOT

Tier = Literal["internal", "bidder"]


class ValidationError(RuntimeError):
    """Raised when a generated governance artifact fails a gate."""


def validate_generated(ssot: GovernanceSSOT, out_dir: Path, tier: Tier) -> None:
    """Validate tier stripping, content hash, and header binding invariants."""

    tier_dir = out_dir / tier
    manifest_path = tier_dir / f"{ssot.package_id}_{tier}_manifest.json"
    xlsx_path = tier_dir / f"{ssot.package_id}_{tier}.xlsx"
    if not manifest_path.exists() or not xlsx_path.exists():
        raise ValidationError(f"missing generated artifacts for tier {tier}")

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    expected_payload = workbook_payload(ssot, tier)
    expected_hash = content_hash(expected_payload)
    if manifest.get("content_hash") != expected_hash:
        raise ValidationError(
            "manifest content hash does not match canonical generated content"
        )

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    expected_sheets = [sheet["name"] for sheet in expected_payload["sheets"]]
    if workbook.sheetnames != expected_sheets:
        raise ValidationError(
            f"workbook sheets drifted: expected {expected_sheets}, observed {workbook.sheetnames}"
        )

    if tier == "bidder":
        _validate_bidder_stripping(ssot, tier_dir, manifest, workbook)

    _validate_header_bindings(ssot)


def _validate_header_bindings(ssot: GovernanceSSOT) -> None:
    for req in ssot.requirements:
        by_name = {binding.header_name: binding for binding in req.header_bindings}
        if len(by_name) != len(req.header_bindings):
            raise ValidationError(f"duplicate header binding names on {req.req_id}")
        for binding in req.header_bindings:
            if binding.observed_index < 1 or not binding.observed_column:
                raise ValidationError(
                    f"invalid observed position for {req.req_id}:{binding.header_name}"
                )


def _validate_bidder_stripping(
    ssot: GovernanceSSOT, tier_dir: Path, manifest: dict[str, object], workbook: object
) -> None:
    forbidden_sheet_names = [
        sheet.name
        for sheet in ssot.generated_sheets
        if sheet.audience in {Audience.internal, Audience.evaluation}
    ]
    leaked_sheets = sorted(set(forbidden_sheet_names) & set(workbook.sheetnames))
    if leaked_sheets:
        raise ValidationError(
            f"bidder workbook leaked non-bidder sheets: {leaked_sheets}"
        )

    manifest_sheet_names = manifest.get("sheet_names", [])
    leaked_manifest_sheets = sorted(
        set(forbidden_sheet_names) & set(manifest_sheet_names)
    )
    if leaked_manifest_sheets:
        raise ValidationError(
            f"bidder manifest leaked non-bidder sheets: {leaked_manifest_sheets}"
        )

    forbidden_terms = set(forbidden_sheet_names)
    for req in ssot.requirements:
        if req.audience in {Audience.internal, Audience.evaluation}:
            forbidden_terms.update({req.req_id, req.title, req.text})

    artifact_text = {
        "workbook": _workbook_text(workbook),
        "html": _read_named_artifact(tier_dir, manifest, "html"),
        "rtm": _read_named_artifact(tier_dir, manifest, "rtm"),
        "manifest": json.dumps(manifest, sort_keys=True),
    }
    for artifact_name, text in artifact_text.items():
        leaked_terms = sorted(term for term in forbidden_terms if term and term in text)
        if leaked_terms:
            raise ValidationError(
                f"bidder {artifact_name} leaked non-bidder content: {leaked_terms}"
            )


def _workbook_text(workbook: object) -> str:
    return "\n".join(
        str(cell)
        for sheet_name in workbook.sheetnames
        for row in workbook[sheet_name].iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )


def _read_named_artifact(
    tier_dir: Path, manifest: dict[str, object], artifact_key: str
) -> str:
    generated_artifacts = manifest.get("generated_artifacts")
    if not isinstance(generated_artifacts, dict):
        raise ValidationError("manifest generated_artifacts must be an object")
    artifact_name = generated_artifacts.get(artifact_key)
    if not isinstance(artifact_name, str):
        raise ValidationError(
            f"manifest missing generated artifact name for {artifact_key}"
        )
    artifact_path = tier_dir / artifact_name
    if not artifact_path.exists():
        raise ValidationError(f"missing generated artifact: {artifact_path}")
    return artifact_path.read_text(encoding="utf-8")
