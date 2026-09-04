"""Generated artifact builder for the ABACUS Contract Governance Workbench."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Literal
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .docx_builder import build_docx
from .io import content_hash
from .pdf_builder import build_pdf
from .pptx_builder import build_pptx
from .schema import Audience, GeneratedSheet, GovernanceSSOT, Requirement

Tier = Literal["internal", "bidder"]


def visible_for_tier(audience: Audience, tier: Tier) -> bool:
    """Return whether an audience classification is visible in an output tier."""

    return tier == "internal" or audience == Audience.bidder


def sheet_rows(
    sheet: GeneratedSheet, ssot: GovernanceSSOT, tier: Tier
) -> list[dict[str, str]]:
    """Build canonical row dictionaries for a generated sheet."""

    requirements = [
        req for req in ssot.requirements if visible_for_tier(req.audience, tier)
    ]
    if sheet.name == "Requirements":
        return [_requirement_row(req) for req in requirements]
    if sheet.name == "Traceability Matrix":
        return [
            {
                "Requirement ID": req.req_id,
                "Source Document": req.source_doc_id,
                "Source Section": req.source_section,
                "Audience": req.audience.value,
            }
            for req in requirements
        ]
    if sheet.name == "Extraction Audit":
        return [
            {
                "Requirement ID": req.req_id,
                "Header Name": binding.header_name,
                "Observed Column": binding.observed_column,
                "Observed Index": str(binding.observed_index),
            }
            for req in requirements
            for binding in req.header_bindings
        ]
    if sheet.name == "Evaluation Notes":
        return [
            {
                "Requirement ID": req.req_id,
                "Audience": req.audience.value,
                "Note": "Internal evaluation placeholder generated from SSOT classification.",
            }
            for req in ssot.requirements
            if req.audience in {Audience.internal, Audience.evaluation}
        ]
    return []


def workbook_payload(ssot: GovernanceSSOT, tier: Tier) -> dict[str, object]:
    """Canonical workbook content used for generation and reproducibility hashing."""

    sheets = []
    for sheet in ssot.generated_sheets:
        if not visible_for_tier(sheet.audience, tier):
            continue
        sheets.append(
            {
                "name": sheet.name,
                "columns": sheet.columns,
                "rows": sheet_rows(sheet, ssot, tier),
            }
        )
    return {"package_id": ssot.package_id, "tier": tier, "sheets": sheets}


def build_artifacts(ssot: GovernanceSSOT, out_dir: Path, tier: Tier) -> dict[str, str]:
    """Generate XLSX, HTML, RTM JSON, DOCX, PPTX, PDF, and a manifest for an output tier."""

    tier_dir = out_dir / tier
    tier_dir.mkdir(parents=True, exist_ok=True)
    payload = workbook_payload(ssot, tier)
    digest = content_hash(payload)

    xlsx_path = tier_dir / f"{ssot.package_id}_{tier}.xlsx"
    html_path = tier_dir / f"{ssot.package_id}_{tier}.html"
    rtm_path = tier_dir / f"{ssot.package_id}_{tier}_rtm.json"
    docx_path = tier_dir / f"{ssot.package_id}_{tier}.docx"
    pptx_path = tier_dir / f"{ssot.package_id}_{tier}.pptx"
    pdf_path = tier_dir / f"{ssot.package_id}_{tier}.pdf"
    manifest_path = tier_dir / f"{ssot.package_id}_{tier}_manifest.json"

    _write_workbook(payload, ssot, xlsx_path)
    _write_html(payload, ssot, html_path)
    _write_json(payload, rtm_path)
    build_docx(payload, ssot, docx_path)
    build_pptx(payload, ssot, pptx_path)
    build_pdf(payload, ssot, pdf_path)
    _write_json(
        {
            "package_id": ssot.package_id,
            "tier": tier,
            "content_hash_algorithm": ssot.build.content_hash_algorithm,
            "content_hash": digest,
            "sheet_names": [sheet["name"] for sheet in payload["sheets"]],
            "generated_artifacts": {
                "xlsx": xlsx_path.name,
                "html": html_path.name,
                "rtm": rtm_path.name,
                "docx": docx_path.name,
                "pptx": pptx_path.name,
                "pdf": pdf_path.name,
            },
        },
        manifest_path,
    )
    return {
        "xlsx": str(xlsx_path),
        "html": str(html_path),
        "rtm": str(rtm_path),
        "docx": str(docx_path),
        "pptx": str(pptx_path),
        "pdf": str(pdf_path),
        "manifest": str(manifest_path),
        "content_hash": digest,
    }


def _requirement_row(req: Requirement) -> dict[str, str]:
    return {
        "Requirement ID": req.req_id,
        "Title": req.title,
        "Requirement Text": req.text,
        "Source Document": req.source_doc_id,
        "Source Section": req.source_section,
        "Audience": req.audience.value,
    }


def _pin_xlsx_core_properties(path: Path, fixed: datetime) -> None:
    """Restore deterministic XLSX core timestamps after openpyxl save-time mutation."""

    core_name = "docProps/core.xml"
    temp_path = path.with_suffix(".tmp.xlsx")
    namespaces = {
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
        "dc": "http://purl.org/dc/elements/1.1/",
        "dcterms": "http://purl.org/dc/terms/",
        "dcmitype": "http://purl.org/dc/dcmitype/",
        "xsi": "http://www.w3.org/2001/XMLSchema-instance",
    }
    for prefix, uri in namespaces.items():
        ET.register_namespace(prefix, uri)

    stamp = fixed.replace(microsecond=0).isoformat(timespec="seconds") + "Z"
    with ZipFile(path, "r") as source, ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            data = source.read(info.filename)
            if info.filename == core_name:
                root = ET.fromstring(data)
                for tag in ("created", "modified"):
                    node = root.find(f"{{{namespaces['dcterms']}}}{tag}")
                    if node is not None:
                        node.text = stamp
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            target.writestr(info, data)
    temp_path.replace(path)


def _write_workbook(
    payload: dict[str, object], ssot: GovernanceSSOT, path: Path
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    fixed = ssot.build.fixed_docprops_timestamp.replace(tzinfo=None)
    wb.properties.created = fixed
    wb.properties.modified = fixed
    wb.properties.creator = "CODEX Contract Governance Generator"
    wb.properties.lastModifiedBy = "CODEX Contract Governance Generator"

    for sheet_payload in payload["sheets"]:  # type: ignore[index]
        ws = wb.create_sheet(sheet_payload["name"])
        columns = sheet_payload["columns"]
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in sheet_payload["rows"]:
            ws.append([row.get(column, "") for column in columns])
        ws.freeze_panes = "A2"
    wb.save(path)
    _pin_xlsx_core_properties(path, fixed)


def _write_html(payload: dict[str, object], ssot: GovernanceSSOT, path: Path) -> None:
    env = Environment(
        loader=FileSystemLoader(Path(__file__).with_name("templates")),
        autoescape=select_autoescape(["html"]),
    )
    template = env.get_template("workbook.html.j2")
    path.write_text(
        template.render(ssot=ssot, payload=payload, content_hash=content_hash(payload)),
        encoding="utf-8",
    )


def _write_json(payload: dict[str, object], path: Path) -> None:
    path.write_text(
        json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
