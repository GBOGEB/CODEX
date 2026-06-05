from __future__ import annotations

import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from codex.contract_governance.builder import build_artifacts
from codex.contract_governance.io import load_ssot
from codex.contract_governance.validator import validate_generated

SSOT = Path("contract_governance/ssot/abacus_contract_governance.yaml")
FIXED_DOC_PROPS = datetime(2026, 1, 1, 0, 0, 0)
GENERATOR = "CODEX Contract Governance Generator"
FORBIDDEN_BIDDER_STRINGS = {
    "Extraction Audit",
    "Evaluation Notes",
    "REQ-W001-001",
    "Generated tier stripping",
    "Internal evaluation placeholder",
}


def test_bootstrap_ssot_declares_bidder_and_internal_tiers() -> None:
    text = SSOT.read_text(encoding="utf-8")

    assert "ABACUS-CGW-W000-W001" in text
    assert "name: Requirements" in text
    assert "name: Extraction Audit" in text
    assert "audience: internal" in text


def test_cli_build_and_validate_execute_end_to_end(tmp_path: Path) -> None:
    build = subprocess.run(
        [
            sys.executable,
            "-m",
            "codex.contract_governance.cli",
            "build",
            "--ssot",
            str(SSOT),
            "--out",
            str(tmp_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    assert "built internal:" in build.stdout
    assert "built bidder:" in build.stdout

    validate = subprocess.run(
        [
            sys.executable,
            "-m",
            "codex.contract_governance.cli",
            "validate",
            "--ssot",
            str(SSOT),
            "--out",
            str(tmp_path),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    assert "validated internal" in validate.stdout
    assert "validated bidder" in validate.stdout


def test_builds_and_validates_both_tiers(tmp_path: Path) -> None:
    ssot = load_ssot(SSOT)
    internal = build_artifacts(ssot, tmp_path, "internal")
    bidder = build_artifacts(ssot, tmp_path, "bidder")

    validate_generated(ssot, tmp_path, "internal")
    validate_generated(ssot, tmp_path, "bidder")

    assert internal["content_hash"] != bidder["content_hash"]
    bidder_wb = load_workbook(bidder["xlsx"], read_only=True, data_only=True)
    assert bidder_wb.sheetnames == ["Requirements", "Traceability Matrix"]


def test_bidder_stripping_is_enforced_in_serialized_artifacts(tmp_path: Path) -> None:
    ssot = load_ssot(SSOT)
    result = build_artifacts(ssot, tmp_path, "bidder")
    validate_generated(ssot, tmp_path, "bidder")

    workbook = load_workbook(result["xlsx"], read_only=True, data_only=True)
    workbook_text = "\n".join(
        str(cell)
        for sheet_name in workbook.sheetnames
        for row in workbook[sheet_name].iter_rows(values_only=True)
        for cell in row
        if cell is not None
    )
    html_text = Path(result["html"]).read_text(encoding="utf-8")
    rtm_text = Path(result["rtm"]).read_text(encoding="utf-8")
    manifest_text = Path(result["manifest"]).read_text(encoding="utf-8")

    for forbidden in FORBIDDEN_BIDDER_STRINGS:
        assert forbidden not in workbook_text
        assert forbidden not in html_text
        assert forbidden not in rtm_text
        assert forbidden not in manifest_text


def test_content_hash_is_manifested_from_canonical_payload(tmp_path: Path) -> None:
    ssot = load_ssot(SSOT)
    result = build_artifacts(ssot, tmp_path, "bidder")
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))

    assert manifest["content_hash_algorithm"] == "sha256"
    assert manifest["content_hash"] == result["content_hash"]
    assert len(manifest["content_hash"]) == 64


def test_repeated_builds_are_content_deterministic_with_pinned_workbook_properties(
    tmp_path: Path,
) -> None:
    ssot = load_ssot(SSOT)
    first = build_artifacts(ssot, tmp_path / "first", "internal")
    second = build_artifacts(ssot, tmp_path / "second", "internal")

    assert first["content_hash"] == second["content_hash"]

    first_manifest = json.loads(Path(first["manifest"]).read_text(encoding="utf-8"))
    second_manifest = json.loads(Path(second["manifest"]).read_text(encoding="utf-8"))
    assert first_manifest["content_hash"] == second_manifest["content_hash"]

    for workbook_path in (first["xlsx"], second["xlsx"]):
        workbook = load_workbook(workbook_path, read_only=True)
        assert workbook.properties.created == FIXED_DOC_PROPS
        assert workbook.properties.modified == FIXED_DOC_PROPS
        assert workbook.properties.creator == GENERATOR
        assert workbook.properties.lastModifiedBy == GENERATOR
